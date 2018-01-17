import tkinter  # assuming Python 3 for simplicity's sake
import tkinter.ttk as ttk
from tkinter import *
from tkinter.ttk import *

root = tkinter.Tk()

from tkinter import *
import sqlite3
import datetime
from tkinter import ttk
from datetime import datetime
#from dateutil.parser import parse
#import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter


class ForestHouse:
    today = datetime.today()
    dd_selection = ''
    mm_selection = ''
    yy_selection = ''
    final_date = ''
    hostdatefield = datetime.today()

    # db_filename = 'C:\\PyProjects\\learnpy\\foresthouse.db'
    db_filename = 'foresthouse.db'

    def __init__(self, root):
        self.root = root
        self.create_gui()

    def execute_db_query(self, query, parameters=()):
        with sqlite3.connect(self.db_filename) as conn:
            cursor = conn.cursor()
            query_result = cursor.execute(query, parameters)
            conn.commit()
        return query_result

    def create_gui(self):
        self.create_left_icon()
        # self.create_label_frame()
        self.create_message_area()
        self.create_tree_view()
        self.create_bottom_buttons()
        self.view_records()
        self.menu_label_frame()

    def create_left_icon(self):
        photo = PhotoImage(file='we1.png')
        label = Label(image=photo)
        label.image = photo
        label.grid(row=0, column=0, sticky='w', padx=140)
        self.root.title("The Foresthouse Guests")

    def create_label_frame(self):
        labelframe = LabelFrame(self.root, text='Add New SQL Record', fg='red')
        labelframe.grid(row=0, column=0, padx=38, pady=8, sticky='w')

        Label(labelframe, text='Name:', relief=RIDGE, width=15).grid(row=1, column=1, sticky=W, pady=2)
        self.namefield = Entry(labelframe, width=45)
        self.namefield.grid(row=1, column=2, sticky=W, padx=5, pady=2)

        Label(labelframe, text='Phone:', relief=RIDGE, width=15).grid(row=2, column=1, sticky=W, pady=2)
        self.phonefield = Entry(labelframe, width=45)
        self.phonefield.grid(row=2, column=2, sticky=W, padx=5, pady=2)

        Label(labelframe, text='Email:', relief=RIDGE, width=15).grid(row=3, column=1, sticky=W, pady=2)
        self.emailfield = Entry(labelframe, width=45)
        self.emailfield.grid(row=3, column=2, sticky=W, padx=5, pady=2)

        Label(labelframe, text='Comment:', relief=RIDGE, width=15).grid(row=4, column=1, sticky=W, pady=2)
        self.commentfield = Entry(labelframe, width=45)
        self.commentfield.grid(row=4, column=2, sticky=W, padx=5, pady=2)

        ttk.Button(labelframe, text=' Save ', command=self.on_add_record_button_clicked).grid(
            row=5, column=2, sticky=W, padx=5, pady=2)

        ttk.Button(labelframe, text='Exit', command=lambda: labelframe.grid_remove()).grid(
            row=5, column=2, sticky=E, padx=5, pady=2)
        self.message['text'] = '                                     '
        self.create_left_icon

    def add_record_xls_label_frame(self):
        labelframe = LabelFrame(self.root, text='Add New XLSX Record', fg='red')
        labelframe.grid(row=0, column=0, padx=8, pady=8, sticky='w')

        Label(labelframe, text='Guest:', relief=RIDGE, width=15).grid(row=1, column=1, sticky=W, padx=10, pady=2)
        self.guestfield = Entry(labelframe, width=45)
        self.guestfield.grid(row=1, column=2, sticky=W, columnspan=3, padx=5, pady=2)
        Label(labelframe, text='Host Date:', relief=RIDGE, width=15).grid(row=2, column=1, sticky=W, padx=10, pady=2)

        dd_options = ("01", "02", "03", "04", "05", "06", "07", "08", "09", "10",
                      "11", "12", "13", "14", "15", "16", "17", "18", "19", "20",
                      "21", "22", "23", "24", "25", "26", "27", "28", "29", "30", "31")
        mm_options = ("01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12")
        yy_options = ("01", "02", "03", "04", "05", "06", "07", "08", "09", "10",
                      "11", "12", "13", "14", "15", "16", "17", "18", "19", "20",
                      "21", "22", "23", "24", "25", "26", "27", "28", "29", "30", "31")
        style = ttk.Style()
        dd_option = StringVar()
        mm_option = StringVar()
        yy_option = StringVar()

        style.configure("TMenubutton", fg="red", background='white')
        # labelframe=labelframe.add_record_xls_label_frame()
        mm_option.set('m')
        dropdwnmm = OptionMenu(labelframe, mm_option, *mm_options)
        dropdwnmm.grid(row=2, column=2, sticky=W, padx=2, pady=2)
        # place(x=150, y=25, height=25)
        dd_option.set('d')
        dropdwndd = OptionMenu(labelframe, dd_option, *dd_options)
        dropdwndd.grid(row=2, column=2, sticky=W, padx=50, pady=2)
        # place(x=200, y=25, height=25)
        yy_option.set('y')
        dropdwnyy = OptionMenu(labelframe, yy_option, *yy_options)
        dropdwnyy.grid(row=2, column=2, sticky=E, padx=5, pady=2)
        # place(x=250, y=25, height=25)

        dd_option.trace('w', lambda x, y, z: callbackdd())
        mm_option.trace('w', lambda x, y, z: callbackmm())
        yy_option.trace('w', lambda x, y, z: callbackyy())

        def callbackdd(*args):
            ForestHouse.dd_selection = dd_option.get()

        def callbackmm(*args):
            ForestHouse.mm_selection = mm_option.get()

        def callbackyy(*args):
            ForestHouse.yy_selection = yy_option.get()
            self.print_final_selection()

        self.print_final_selection()
        self.hostdatefield = ForestHouse.final_date

        Label(labelframe, text='# Nights:', relief=RIDGE, width=8).grid(row=2, column=3, sticky=E, padx=3, pady=2)
        self.nitsfield = Entry(labelframe, width=8)
        self.nitsfield.grid(row=2, column=4, sticky=E, padx=5, pady=2)

        Label(labelframe, text='Price:', relief=RIDGE, width=15).grid(row=3, column=1, sticky=W, padx=10, pady=2)
        self.pricefield = Entry(labelframe, width=15)
        self.pricefield.grid(row=3, column=2, sticky=W, padx=5, pady=2)

        Label(labelframe, text='Cleaning:', relief=RIDGE, width=8).grid(row=3, column=3, sticky=E, padx=3, pady=2)
        self.cleaningfield = Entry(labelframe, width=8)
        self.cleaningfield.grid(row=3, column=4, sticky=E, padx=5, pady=2)

        Label(labelframe, text='Airbnb Fee:', relief=RIDGE, width=15).grid(row=4, column=1, sticky=W, padx=10, pady=2)
        self.feefield = Entry(labelframe, width=15)
        self.feefield.grid(row=4, column=2, sticky=W, padx=5, pady=2)

        Label(labelframe, text='Paid:', relief=RIDGE, width=8).grid(row=4, column=3, sticky=E, padx=3, pady=2)
        self.paidfield = Entry(labelframe, width=8)
        self.paidfield.grid(row=4, column=4, sticky=E, padx=5, pady=2)

        ttk.Button(labelframe, text='Save', command=self.on_add_xls_record_button_clicked).grid(
            row=5, column=1, sticky=W, padx=11, pady=2)

        ttk.Button(labelframe, text='Exit', command=lambda: labelframe.grid_remove()).grid(
            row=5, column=2, sticky=W, padx=3, pady=2)
        self.message['text'] = '                                     '
        self.create_left_icon

    def print_final_selection(self):
        if ForestHouse.mm_selection == '' or ForestHouse.dd_selection == '' or ForestHouse.yy_selection == '':
            ForestHouse.final_date = datetime.today()
        else:
            ForestHouse.final_date = ForestHouse.mm_selection + '-' + ForestHouse.dd_selection + '-20' + ForestHouse.yy_selection
        if len(str(ForestHouse.final_date)) == 10:
            datetime.strptime(ForestHouse.final_date, '%m-%d-%Y')

    def menu_label_frame(self):
        labelframe2 = LabelFrame(self.root, text='Foresthouse Menu')
        labelframe2.grid(row=0, column=0, padx=40, pady=0, sticky='e')

        ttk.Button(labelframe2, text='Accounting', width=40, command=self.on_accounting_button_clicked).grid(
            row=1, column=1, sticky=W, padx=2, pady=2)

        ttk.Button(labelframe2, text='Activities', width=40, command=self.on_activities_button_clicked).grid(
            row=2, column=1, sticky=W, padx=2, pady=1)
        ttk.Button(labelframe2, text='Calculator', width=40, command=self.on_calculator_button_clicked).grid(
            row=3, column=1, sticky=W, padx=2, pady=1)

        ttk.Button(labelframe2, text='CalcTax', width=40, command=self.on_calctax_button_clicked).grid(
            row=4, column=1, sticky=W, padx=2, pady=1)
        ttk.Button(labelframe2, text='Contacts', width=40, command=self.on_contats_button_clicked).grid(
            row=5, column=1, sticky=W, padx=2, pady=1)

    def menu_label_frame2(self):
        labelframe2 = LabelFrame(self.root, text='Foresthouse Menu')
        labelframe2.grid(row=0, column=0, padx=40, pady=0, sticky='e')

        ttk.Button(labelframe2, text='SQLite DataBase', width=40,
                   command=self.on_exit_accounting_selected_button_clicked).grid(
            row=1, column=1, sticky=W, padx=2, pady=2)

        ttk.Button(labelframe2, text='Activities', width=40, command=self.on_activities_button_clicked).grid(
            row=2, column=1, sticky=W, padx=2, pady=1)
        ttk.Button(labelframe2, text='Calculator', width=40, command=self.on_calculator_button_clicked).grid(
            row=3, column=1, sticky=W, padx=2, pady=1)

        ttk.Button(labelframe2, text='CalcTax', width=40, command=self.on_calctax_button_clicked).grid(
            row=4, column=1, sticky=W, padx=2, pady=1)
        ttk.Button(labelframe2, text='Contacts', width=40, command=self.on_contats_button_clicked).grid(
            row=5, column=1, sticky=W, padx=2, pady=1)

    def on_add_record_button_clicked(self):
        self.add_new_record()
        self.message['text'] = '                                     '
        self.create_left_icon

    def on_add_xls_record_button_clicked(self):
        self.add_new_xls_record()
        self.message['text'] = '                                     '
        self.create_left_icon

    def on_calctax_button_clicked(self):
        import subprocess
        subprocess.call(["python3", "calctax.py"])

    def on_accounting_button_clicked(self):
        self.menu_label_frame2()
        self.create_xls_tree_view()

    def on_calculator_button_clicked(self):
        import subprocess
        subprocess.call(["python3", "calculator.py"])
        # subprocess.call(["python", "myscript2.py"])

    def on_contats_button_clicked(self):
        import subprocess
        subprocess.call(["python3", "contacts.py"])

    def on_activities_button_clicked(self):
        import subprocess
        subprocess.call(["python3", "api2.py"])

        # self.create_activities_xls_tree_view()

    def create_tree_view(self):
        self.tree = ttk.Treeview()

        self.tree = ttk.Treeview(height=14, columns=1)
        self.tree.grid(row=4, column=0)

        self.tree.heading('#0', text='ID', anchor=W)
        self.tree.column("#0", width=40, stretch=NO)
        style = ttk.Style()

        style.configure("Treeview.Heading", background='blue', foreground='blue')
        self.tree["columns"] = ("one", "two", "three", "four")
        self.tree.column("one", width=150, stretch=NO)
        self.tree.column("two", width=100, stretch=NO)
        self.tree.column("three", width=200, stretch=NO)
        self.tree.column("four", width=250, stretch=NO)

        self.tree.heading("one", text="Name", anchor=W)
        self.tree.heading("two", text="Phone", anchor=W)
        self.tree.heading("three", text="Email", anchor=W)
        self.tree.heading("four", text="Comments", anchor=W)

    def create_xls_tree_view(self):
        self.xlstree = ttk.Treeview()

        self.xlstree = ttk.Treeview(height=14, columns=0)
        self.xlstree.grid(row=4, column=0)
        self.xlstree.heading('#0')
        # self.xlstree.column("#0", width=0)
        self.xlstree.column("#0", width=0, stretch=NO)

        style = ttk.Style()

        style.configure("Treeview.Heading", background='blue', foreground='blue')
        self.xlstree["columns"] = ("", "one", "two", "three", "four", "five", "six", "seven")

        self.xlstree.column("one", width=67, stretch=NO)
        self.xlstree.column("two", width=50, stretch=NO)
        self.xlstree.column("three", width=60, stretch=NO)
        self.xlstree.column("four", width=60, stretch=NO)
        self.xlstree.column("five", width=60, stretch=NO)
        self.xlstree.column("six", width=90, stretch=NO)
        self.xlstree.column("seven", width=155, stretch=NO)

        self.xlstree.heading("", text="Guest/Vendor", anchor=W)
        self.xlstree.heading("one", text="Date", anchor=W, )
        self.xlstree.heading("two", text="Nights", anchor=W)
        self.xlstree.heading("three", text="Price", anchor=W)
        self.xlstree.heading("four", text="Cleaning", anchor=W)
        self.xlstree.heading("five", text="Fees", anchor=W)
        self.xlstree.heading("six", text="Paid", anchor=W)
        self.xlstree.heading("seven", text="Comment", anchor=W)

        wb = load_workbook(filename='airbnb.xlsx')
        ws = wb.active
        row_count = ws.max_row
        column_count = ws.max_column
        items = self.xlstree.get_children()

        for item in items:
            self.xlstree.delete(item)
        for row in ws.iter_rows(min_row=1, min_col=1, max_col=8):
            for cell in row:
                if cell.value in ("Guest", "Host Date", "Nts", "Price", "Cleaning Fee", "Fee", "Paid", "Comment"):
                    break
                else:
                    self.xlstree.insert('', 'end', values=[cell.value for cell in row])
                    break

        self.xlstree.bind('<ButtonRelease-1>', self.select_xlsx_item)

        self.create_xls_bottom_buttons()

    def select_item(self, a):  # added self and a (event)
        test_str_library = self.tree.item(self.tree.selection())
        item = self.tree.selection()  # [0]  # which row did you click on
        record_id = self.tree.item(item)['text']

    def select_xls_item(self, event):
        for item in self.xlstree.selection():
            item_text = self.xlstree.item(item, "text")

    def select_xlsx_item(self, a):  # added self and a (event)
        # gets all the values of the selected row
        test_str_library = self.xlstree.item(self.xlstree.selection())
        item = self.xlstree.selection()  # [0]  # which row did you click on
        guestfield = self.xlstree.item(item)['values'][0]

        return guestfield

    def on_delete_selected_button_clicked(self):
        self.message['text'] = ''
        try:
            self.tree.item(self.tree.selection())['values'][0]
        except IndexError as e:
            self.message['text'] = 'No item selected to delete     '
            return

        self.delete_record()

    def on_delete_xls_selected_button_clicked(self):
        self.message['text'] = ''
        try:
            self.xlstree.item(self.xlstree.selection())['values'][0]
        except IndexError as e:
            self.message['text'] = 'No item selected to delete     '
            return

        #self.delete_xls_record()
        self.del_xls_row()

    def del_xls_row(self):
        import openpyxl
        wb = openpyxl.load_workbook('airbnb.xlsx')
        ws1 = wb.get_sheet_by_name('Sheet1')
        wb.create_sheet("Sheet2")
        ws2 = wb.get_sheet_by_name('Sheet2')

        item = self.xlstree.selection() #[0]
        guestfield = self.xlstree.item(item)['values'][0]
        # convert hex to base 10
        row_to_delete = int(str(item)[3:6], 16)+1

        print("row to delete is:",row_to_delete)

        cell1='A'+str(row_to_delete)

        print(cell1)
        # If you want to Start at Row 2 to append Row Data
        # Set Private self._current_row to 1
        #ws2.cell(row=1, column=1).value = ws2.cell(row=1, column=1).value

        # Define min/max Column Range to copy
        from openpyxl.utils import range_boundaries
        min_col, min_row, max_col, max_row = range_boundaries('A:GH')

        # Define Cell Index (0 Based) used to Check Value
        check = 0  # == A

        for row in ws1.iter_rows():
            if row[check].value == 'TrueValue':
                # Copy Row Values
                # We deal with Tuple Index 0 Based, so min_col must have to be -1
                ws2.append((cell.value for cell in row[min_col-1:max_col]))

        wb.save('airbnb.xlsx')

        self.create_xls_tree_view()

    def delete_record(self):
        item = self.tree.selection()  # [0]
        idfield = self.tree.item(item)['text']
        conn = sqlite3.connect(self.db_filename)
        c = conn.cursor()
        conn.text_factory = str
        data3 = str(idfield)
        query = "DELETE FROM guest WHERE id = '%s';" % data3
        mydata = c.execute(query)
        conn.commit()
        self.view_records()

    def delete_xls_record(self):
        item = self.xlstree.selection()  # [0]
        guestfield = self.xlstree.item(item)['values'][0]
        # convert hex to base 10
        row_to_delete = int(str(item)[3:6], 16) + 1

        print("row to delete is:", row_to_delete)

        import win32com.client

        wb = load_workbook(filename='airbnb.xlsx')
        ws = wb.get_sheet_by_name('Sheet1')
        ws = wb.active
        mycell = ws.cell(row=row_to_delete, column=1)
        value1 = str(mycell.value)
        value1 = value1[0:7]
        if value1 != 'Deleted' or value1 == '' or value1 == None:
            mycell.value = 'Deleted -' + str(mycell.value)

        ws = wb.get_sheet_by_name('Sheet1')
        ws = wb.active

        wb.copy_worksheet(ws)
        # page4 = wb.get_sheet_by_name('Sheet1 Copy')
        # page4.title = 'page4'
        ws = wb.get_sheet_by_name('Sheet1')
        ws = wb.active
        wb.save('airbnb.xlsx')
        self.create_xls_tree_view()

    def on_modify_selected_button_clicked(self):
        self.message['text'] = ''
        try:
            self.tree.item(self.tree.selection())['values'][0]
        except IndexError as e:
            self.message['text'] = 'No item selected to modify     '
            return
        self.open_modify_window()

    def on_modify_xls_selected_button_clicked(self):
        self.message['text'] = ''
        try:
            self.xlstree.item(self.xlstree.selection())['values'][0]
        except IndexError as e:
            self.message['text'] = 'No item selected to modify     '
            return
        self.open_modify_xls_window()

    def on_print_selected_button_clicked(self):
        idfield = self.tree.item(self.tree.selection())['text']
        namefield = self.tree.item(self.tree.selection())['values'][0]
        phonefield = self.tree.item(self.tree.selection())['values'][1]
        emailfield = self.tree.item(self.tree.selection())['values'][2]
        commentfield = self.tree.item(self.tree.selection())['values'][3]
        import win32ui
        import win32print
        # X from the left margin, Y from top margin both in pixels
        X = 50
        Y = 50
        multi_line_string = 'Mike'
        hDC = win32ui.CreateDC()
        hDC.CreatePrinterDC(win32print.GetDefaultPrinter())
        hDC.StartDoc('Microsoft Print to PDF')
        hDC.StartPage()
        item = self.tree.selection()[0]  # which row did you click on
        multi_line_string = (
        str(idfield) + '  ' + namefield + '  ' + phonefield + '  ' + emailfield + '  ' + commentfield)

        hDC.TextOut(0, 0, multi_line_string)
        hDC.EndPage()
        hDC.EndDoc()

    def on_print_xls_selected_button_clicked(self):
        item = self.xlstree.selection()  # [0]

        guestfield = self.xlstree.item(self.xlstree.selection())['values'][0]
        hostdatefield = self.xlstree.item(self.xlstree.selection())['values'][1]
        nitsfield = self.xlstree.item(self.xlstree.selection())['values'][2]
        pricefield = self.xlstree.item(self.xlstree.selection())['values'][3]
        cleaningfield = self.xlstree.item(self.xlstree.selection())['values'][4]
        feefield = self.xlstree.item(self.xlstree.selection())['values'][5]
        paidfield = self.xlstree.item(self.xlstree.selection())['values'][6]
        commentfield = self.xlstree.item(self.xlstree.selection())['values'][7]

        import win32ui
        import win32print
        # X from the left margin, Y from top margin both in pixels
        X = 50
        Y = 50
        multi_line_string = 'Mike'
        hDC = win32ui.CreateDC()
        hDC.CreatePrinterDC(win32print.GetDefaultPrinter())
        hDC.StartDoc('Microsoft Print to PDF')
        hDC.StartPage()
        item = self.tree.selection()[0]  # which row did you click on
        multi_line_string = (
        guestfield + '  ' + str(hostdatefield) + '  ' + str(nitsfield) + '  ' + str(pricefield) + '  ' +
        str(cleaningfield) + '  ' + str(feefield) + '  ' + str(paidfield) + '  ' + commentfield)

        hDC.TextOut(0, 0, multi_line_string)
        hDC.EndPage()
        hDC.EndDoc()

    def open_modify_xls_window(self):

        # self.hostdatefield = ForestHouse.final_date
        item = self.xlstree.selection()  # [0]

        guestfield = self.xlstree.item(self.xlstree.selection())['values'][0]
        hostdatefield = self.xlstree.item(self.xlstree.selection())['values'][1]
        nitsfield = self.xlstree.item(self.xlstree.selection())['values'][2]
        pricefield = self.xlstree.item(self.xlstree.selection())['values'][3]
        cleaningfield = self.xlstree.item(self.xlstree.selection())['values'][4]
        feefield = self.xlstree.item(self.xlstree.selection())['values'][5]
        paidfield = self.xlstree.item(self.xlstree.selection())['values'][6]
        commentfield = self.xlstree.item(self.xlstree.selection())['values'][7]

        labelframe1 = LabelFrame(self.root, text='Modify XLSX Record', fg='red')
        labelframe1.grid(row=0, column=0, padx=8, pady=8, sticky='w')

        Label(labelframe1, text='Guest:', relief=RIDGE, width=15).grid(row=1, column=1, sticky=W, padx=10, pady=2)
        guestfield1 = Entry(labelframe1, width=44)
        guestfield1.insert(44, guestfield)
        guestfield1.grid(row=1, column=2, sticky=W, columnspan=3, padx=5, pady=2)
        guestfield = guestfield1.get()

        Label(labelframe1, text='Host Date:', relief=RIDGE, width=15).grid(row=2, column=1, sticky=W, padx=10, pady=2)
        hostdatefield1 = Entry(labelframe1, width=20)
        hostdatefield1.insert(20, hostdatefield)
        hostdatefield1.grid(row=2, column=2, sticky=W, columnspan=1, padx=5, pady=2)
        hostdatefield = hostdatefield1.get()

        Label(labelframe1, text='# Nights:', relief=RIDGE, width=8).grid(row=2, column=3, sticky=E, padx=3, pady=2)
        nitsfield1 = Entry(labelframe1, width=8)
        nitsfield1.insert(8, nitsfield)
        nitsfield1.grid(row=2, column=4, sticky=W, padx=8, pady=2)
        nitsfield = nitsfield1.get()

        Label(labelframe1, text='Price:', relief=RIDGE, width=15).grid(row=3, column=1, sticky=W, padx=10, pady=2)
        pricefield1 = Entry(labelframe1, width=15)
        pricefield1.insert(15, pricefield)
        pricefield1.grid(row=3, column=2, sticky=W, padx=5, pady=2)
        pricefield = pricefield1.get()

        Label(labelframe1, text='Cleaning:', relief=RIDGE, width=8).grid(row=3, column=3, sticky=E, padx=3, pady=2)
        cleaningfield1 = Entry(labelframe1, width=8)
        cleaningfield1.insert(8, cleaningfield)
        cleaningfield1.grid(row=3, column=4, sticky=W, padx=8, pady=2)

        Label(labelframe1, text='Airbnb Fee:', relief=RIDGE, width=15).grid(row=4, column=1, sticky=W, padx=10, pady=2)
        feefield1 = Entry(labelframe1, width=15)
        feefield1.insert(15, feefield)
        feefield1.grid(row=4, column=2, sticky=W, padx=5, pady=2)

        Label(labelframe1, text='Paid:', relief=RIDGE, width=8).grid(row=4, column=3, sticky=E, padx=3, pady=2)
        paidfield1 = Entry(labelframe1, width=8)
        paidfield1.insert(8, paidfield)
        paidfield1.grid(row=4, column=4, sticky=W, padx=8, pady=2)

        Label(labelframe1, text='Comment:', relief=RIDGE, width=15).grid(row=5, column=1, sticky=W, padx=10, pady=2)
        commentfield1 = Entry(labelframe1, width=20)
        commentfield1.insert(20, commentfield)
        commentfield1.grid(row=5, column=2, sticky=W, columnspan=1, padx=5, pady=2)
        commentfield = commentfield1.get()

        b1 = Button(labelframe1, text='Save', fg='red', command=lambda: self.on_update_xls_record_button_clicked(
            guestfield1.get(), hostdatefield1.get(), nitsfield1.get(), pricefield1.get(),
            cleaningfield1.get(), feefield1.get(), paidfield1.get(), commentfield1.get()))
        b1.grid(row=5, column=3, sticky=E, padx=1, pady=2)

        b2 = Button(labelframe1, text='  Exit  ', fg='red', command=lambda: labelframe1.grid_remove())
        b2.grid(row=5, column=4, sticky=W, padx=8, pady=2)

        self.message['text'] = '                                     '
        self.create_left_icon

    def on_update_xls_record_button_clicked(self, guestfield1, hostdatefield1, nitsfield1, pricefield1,
                                            cleaningfield1, feefield1, paidfield1, commentfield1):

        item = self.xlstree.selection()  # [0]
        guestfield = self.xlstree.item(item)['values'][0]

        # convert hex to base 10
        row_to_update = int(str(item)[3:6], 16) + 1

        import win32com.client

        wb = load_workbook(filename='airbnb.xlsx')
        ws = wb.get_sheet_by_name('Sheet1')
        ws = wb.active
        mycell1 = ws.cell(row=row_to_update, column=1)
        mycell2 = ws.cell(row=row_to_update, column=2)
        mycell3 = ws.cell(row=row_to_update, column=3)
        mycell4 = ws.cell(row=row_to_update, column=4)
        mycell5 = ws.cell(row=row_to_update, column=5)
        mycell6 = ws.cell(row=row_to_update, column=6)
        mycell7 = ws.cell(row=row_to_update, column=7)
        mycell8 = ws.cell(row=row_to_update, column=8)

        mycell1.value = guestfield1
        mycell2.value = hostdatefield1
        if nitsfield1 not in (None, 'None', ''):
            mycell3.value = int(nitsfield1)
        if pricefield1 not in (None, 'None', ''):
            mycell4.value = round(float(pricefield1), 2)
        if cleaningfield1 not in (None, 'None', ''):
            mycell5.value = round(float(cleaningfield1), 2)
        if feefield1 not in (None, 'None', ''):
            mycell6.value = round(float(feefield1), 2)
        if paidfield1 not in (None, 'None', ''):
            mycell7.value = round(float(paidfield1), 2)

        mycell8.value = commentfield1

        # value1 = str(mycell1.value)
        # value1 = value1[0:7]
        # if value1 != 'Deleted' or value1 == '' or value1 == None:
        #   mycell.value = 'Deleted -' + str(mycell.value)
        # ws = wb.get_sheet_by_name('Sheet1')
        # ws = wb.active
        # wb.copy_worksheet(ws)
        # page4 = wb.get_sheet_by_name('Sheet1 Copy')
        # page4.title = 'page4'

        ws = wb.get_sheet_by_name('Sheet1')
        ws = wb.active
        wb.save('airbnb.xlsx')

        self.create_xls_tree_view()

    def open_modify_window(self):
        win = Tk()
        Label(win, text="Name").grid(row=0)
        Label(win, text="phone").grid(row=1)
        Label(win, text="Email").grid(row=2)
        Label(win, text="Comment").grid(row=3)

        item = self.tree.selection()  # [0]
        idfield = self.tree.item(item)['text']

        namefield = self.tree.item(self.tree.selection())['values'][0]
        phonefield = self.tree.item(self.tree.selection())['values'][1]
        emailfield = self.tree.item(self.tree.selection())['values'][2]
        commentfield = self.tree.item(self.tree.selection())['values'][3]

        e1 = Entry(win, width=30)
        e2 = Entry(win, width=30)
        e3 = Entry(win, width=30)
        e4 = Entry(win, width=30)

        e1.insert(20, namefield)
        e2.insert(20, phonefield)
        e3.insert(20, emailfield)
        e4.insert(20, commentfield)

        e1.grid(row=0, column=1, padx=10, pady=5)
        e2.grid(row=1, column=1, padx=10, pady=5)
        e3.grid(row=2, column=1, padx=10, pady=5)
        e4.grid(row=3, column=1, padx=10, pady=5)

        Button(win, text='Quit', command=win.destroy, fg='green').grid(row=4, column=1, sticky=W, pady=4, padx=10)
        Button(win, text='Update', fg='red', command=lambda:
        self.modify_record(e1.get(), e2.get(), e3.get(), e4.get())).grid(row=4, column=1, sticky=E, pady=4, padx=10)

    def modify_record(self, namefield, phonefield, emailfield, commentfield):
        item = self.tree.selection()  # [0]
        idfield = self.tree.item(item)['text']

        query = 'UPDATE guest SET name=?, phone=?, email=?, comment=? WHERE  id=?'
        parameters = (namefield, phonefield, emailfield, commentfield, idfield)

        with sqlite3.connect(self.db_filename) as conn:
            cursor = conn.cursor()
            query_result = cursor.execute(query, parameters)
            conn.commit()
        self.message['text'] = 'The record of  {}  was modified     '.format(namefield)
        self.view_records()
        return

    def add_update_record(self):
        if self.new_records_validated():
            query = 'INSERT INTO guest VALUES(NULL,?, ?, ?, ?)'
            parameters = (self.namefield.get(), self.phonefield.get(),
                          self.emailfield.get(), self.commentfield.get())
            self.execute_db_query(query, parameters)
            self.message['text'] = 'Phone record of {} {} {} {} added     '.format(
                self.namefield.get(), self.phonefield.get(), self.emailfield.get(), self.commentfield.get())
            self.namefield.delete(0, END)
            self.phonefield.delete(0, END)
            self.emailfield.delete(0, END)
            self.commentfield.delete(0, END)
        else:
            self.message['text'] = 'Name cannot be blank     '
        self.view_records()

    def add_new_record(self):
        if self.new_records_validated():
            query = 'INSERT INTO guest VALUES(NULL,?, ?, ?, ?)'
            parameters = (self.namefield.get(), self.phonefield.get(),
                          self.emailfield.get(), self.commentfield.get())
            self.execute_db_query(query, parameters)
            self.message['text'] = 'New SQL record of {} added     '.format(self.namefield.get())
            self.namefield.delete(0, END)
            self.phonefield.delete(0, END)
            self.emailfield.delete(0, END)
            self.commentfield.delete(0, END)
        else:
            self.message['text'] = 'Name cannot be blank     '

        self.view_records()

    def add_new_xls_record(self):
        class Foo(float):
            def __new__(cls, value, extra):
                return super().__new__(cls, value)

            def __init__(self, value, extra):
                float.__init__(value)
                self.extra = extra

        if self.new_xls_records_validated():
            style = ttk.Style()
            from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
            parameters = (self.guestfield.get(), ForestHouse.final_date,
                          self.nitsfield.get(), self.pricefield.get(), self.cleaningfield.get(),
                          self.feefield.get(), self.paidfield.get())
            item = self.tree.selection()  # [0]

            guestfield = self.guestfield.get()

            #########################################
            # from datetime import datetime
            # from dateutil.parser import parse
            # import pandas as pd
            # war_start = '2011-01-03'
            # datetime.strptime(war_start, '%Y-%m-%d')
            ###########################################

            hostdatefield = ForestHouse.final_date

            if self.nitsfield.get() != '':
                nitsfield = int(self.nitsfield.get())
            else:
                nitsfield = self.nitsfield.get()

            if self.pricefield.get() != '':
                pricefield = float(self.pricefield.get())
            else:
                pricefield = self.pricefield.get()

            if self.cleaningfield.get() != '':
                cleaningfield = float(self.cleaningfield.get())
            else:
                cleaningfield = self.cleaningfield.get()

            if self.feefield.get() != '':
                feefield = float(self.feefield.get())
            else:
                feefield = self.feefield.get()

            if self.paidfield.get() != '':
                paidfield = float(self.paidfield.get())
            else:
                paidfield = self.paidfield.get()

            wb = load_workbook(filename='airbnb.xlsx')
            ws = wb.active
            row_count = ws.max_row + 1
            column_count = ws.max_column
            ws.cell(row=row_count, column=1).value = guestfield
            ws.cell(row=row_count, column=2).value = hostdatefield
            ws.cell(row=row_count, column=3).value = nitsfield
            ws.cell(row=row_count, column=4).value = pricefield
            ws.cell(row=row_count, column=5).value = cleaningfield
            ws.cell(row=row_count, column=6).value = feefield
            ws.cell(row=row_count, column=7).value = paidfield
            wb.save(filename='airbnb.xlsx')

            self.create_xls_tree_view()
            self.message['text'] = 'New XLS record of {} added     '.format(self.guestfield.get())
            self.guestfield.delete(0, END)
            # self.hostdatefield.delete(0, END)
            self.nitsfield.delete(0, END)
            self.pricefield.delete(0, END)
            self.cleaningfield.delete(0, END)
            self.feefield.delete(0, END)
            self.paidfield.delete(0, END)

        else:
            self.message['text'] = 'Name cannot be blank     '

            # self.message['text'] = ''
            # self.view_xls_records()

    def create_bottom_buttons(self):
        ttk.Button(text='Add Record',
                   command=self.create_label_frame).grid(row=5, column=0, sticky=W, padx=15, pady=5)  # W5
        ttk.Button(text='Modify Selected',
                   command=self.on_modify_selected_button_clicked).grid(row=5, column=0, sticky=W, padx=200,
                                                                        pady=5)  # W200
        ttk.Button(text='Delete Selected',
                   command=self.on_delete_selected_button_clicked).grid(row=5, column=0, sticky=E, padx=200,
                                                                        pady=5)  # E200
        ttk.Button(text='Print Records',
                   command=self.on_print_selected_button_clicked).grid(row=5, column=0, sticky=E, padx=15, pady=5)  # E5

    def clear_button(self):
        for button in root.grid_slaves():
            if int(button.grid_info()["row"]) > 4:
                button.grid_forget()

    def create_xls_bottom_buttons(self):
        self.clear_button()

        ttk.Button(text='Add Record',
                   command=self.add_record_xls_label_frame).grid(row=5, column=0, sticky=W, padx=15, pady=5)
        ttk.Button(text='Modify Selected',
                   command=self.on_modify_xls_selected_button_clicked).grid(row=5, column=0, sticky=W, padx=150, pady=5)
        ttk.Button(text='Delete Selected',
                   command=self.on_delete_xls_selected_button_clicked).grid(row=5, column=0, sticky=W, padx=320, pady=5)
        ttk.Button(text='Print Records',
                   command=self.on_print_xls_selected_button_clicked).grid(row=5, column=0, sticky=E, padx=175, pady=5)
        #ttk.Button(text='Remove Deleted',
        #           command=self.del_xls_row).grid(row=5, column=0, sticky=E, padx=15)

    def on_exit_accounting_selected_button_clicked(self):
        self.menu_label_frame()
        self.message['text'] = '                                     '
        self.clear_button()
        self.create_tree_view()
        self.create_bottom_buttons()
        self.view_records()

    def view_records(self):
        items = self.tree.get_children()
        for item in items:
            self.tree.delete(item)
        query = 'SELECT * FROM guest ORDER BY ID asc'
        guest_table_entries = self.execute_db_query(query)

        for row in guest_table_entries:
            self.tree.insert('', 'end', text=row[0], values=(row[1], row[2], row[3], row[4]))
        self.tree.bind('<ButtonRelease-1>', self.select_item)

    def view_xls_records(self):
        wb = load_workbook(filename='airbnb.xlsx')
        ws = wb.active
        row_count = ws.max_row
        column_count = ws.max_column
        # self.tree = ttk.Treeview()
        items = self.tree.get_children()
        for item in items:
            self.tree.delete(item)

        for row in ws.iter_rows(min_row=1, min_col=1, max_col=7):
            for cell in row:
                if cell.value in ("Guest", "Host Date", "Nts", "Price", "Cleaning Fee", "Fee", "Paid", "Comment"):
                    break
                else:
                    self.tree.insert('', 'end', values=[cell.value for cell in row])
                    break

        self.tree.bind('<ButtonRelease-1>', self.select_xls_item)

    def new_records_validated(self):
        return (self.namefield.get() != 0 or ForestHouse.hostdatefield != 0)  # and len(self.phonefield.get()) != 0

    def new_xls_records_validated(self):
        return (self.guestfield.get() != 0 or ForestHouse.hostdatefield != 0)  # and len(self.phonefield.get()) != 0

    def create_message_area(self):
        self.message = Label(text='', fg='red')
        self.message.grid(row=3, column=0, sticky=E)


if __name__ == '__main__':
    # root = Tk()


    # def resize(event):
    #    print("New size is: {}x{}".format(event.width, event.height))
    # root.bind("<Configure>", resize)

    # w, h = root.winfo_screenwidth(), root.winfo_screenheight()
    # root.geometry("%dx%d+0+0" % (w,h))

    # root.geometry("%dx%d+0+0" % (743, 529))
    root.geometry("%dx%d+50+50" % (743, 529))
    '''
    RTitle = root.title("Forest House System")
    RWidth = root.winfo_screenwidth()
    RHeight = root.winfo_screenheight()
    root.geometry(("%dx%d") % (RWidth, RHeight))
    '''
    application = ForestHouse(root)
    root.mainloop()
