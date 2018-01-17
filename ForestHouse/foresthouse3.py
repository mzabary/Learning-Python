from tkinter import *
from tkinter import ttk
import sqlite3

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter

class PhoneBook:
    # db_filename = 'C:\\PyProjects\\learnpy\\foresthouse.db'
    db_filename = 'foresthouse.db'

    def __init__(self, root):
        self.root = root
        self.create_gui()

    def execute_db_query(self, query, parameters=()):
        with sqlite3.connect(self.db_filename) as conn:
            cursor = conn.cursor()
            query_result = cursor.execute(query, parameters)
            conn.commit()
        return query_result

    def create_gui(self):
        self.create_left_icon()
        # self.create_label_frame()
        self.create_message_area()
        self.create_tree_view()
        self.create_bottom_buttons()
        self.view_records()

        self.menu_label_frame()
        # self.select_item()

    def create_left_icon(self):
        # photo = PhotoImage(file='phonebook.gif')
        # photo = PhotoImage(file='C:\\PyProjects\\learnpy\\we1.png')
        photo = PhotoImage(file='we1.png')
        label = Label(image=photo)
        label.image = photo
        label.grid(row=0, column=0, sticky='w', padx=140)
        self.root.title("The Foresthouse Guests")

    def create_label_frame(self):
        labelframe = LabelFrame(self.root, text='Add New SQL Record',fg='red')
        labelframe.grid(row=0, column=0, padx=38, pady=8, sticky='w')

        Label(labelframe, text='Name:', relief=RIDGE, width=15).grid(row=1, column=1, sticky=W, pady=2)
        self.namefield = Entry(labelframe, width=40)
        self.namefield.grid(row=1, column=2, sticky=W, padx=5, pady=2)

        Label(labelframe, text='Phone:', relief=RIDGE, width=15).grid(row=2, column=1, sticky=W, pady=2)
        self.phonefield = Entry(labelframe, width=40)
        self.phonefield.grid(row=2, column=2, sticky=W, padx=5, pady=2)

        Label(labelframe, text='Email:', relief=RIDGE, width=15).grid(row=3, column=1, sticky=W, pady=2)
        self.emailfield = Entry(labelframe, width=40)
        self.emailfield.grid(row=3, column=2, sticky=W, padx=5, pady=2)

        Label(labelframe, text='Comment:', relief=RIDGE, width=15).grid(row=4, column=1, sticky=W, pady=2)
        self.commentfield = Entry(labelframe, width=40)
        self.commentfield.grid(row=4, column=2, sticky=W, padx=5, pady=2)

        ttk.Button(labelframe, text=' Save ', command=self.on_add_record_button_clicked).grid(
            row=5, column=2, sticky=W, padx=5, pady=2)

        ttk.Button(labelframe, text='Exit', command=lambda: labelframe.grid_remove()).grid(
            row=5, column=2, sticky=E, padx=5, pady=2)

        self.create_left_icon

    def add_record_xls_label_frame(self):
        labelframe = LabelFrame(self.root, text='Add New XLSX Record',fg='red')
        labelframe.grid(row=0, column=0, padx=8, pady=8, sticky='w')

        Label(labelframe, text='Guest:', relief=RIDGE, width=15).grid(row=1, column=1, sticky=W, padx=10,pady=2)
        self.guestfield = Entry(labelframe, width=45)
        self.guestfield.grid(row=1, column=2, sticky=W, columnspan=3,padx=5, pady=2)

        Label(labelframe, text='Host Date:', relief=RIDGE, width=15).grid(row=2, column=1, sticky=W,padx=10, pady=2)
        self.hostdatefield = Entry(labelframe, width=15)
        self.hostdatefield.grid(row=2, column=2, sticky=W, padx=5, pady=2)

        Label(labelframe, text='# Nights:', relief=RIDGE, width=8).grid(row=2, column=3, sticky=E, padx=3,pady=2)
        self.nitsfield = Entry(labelframe, width=8)
        self.nitsfield.grid(row=2, column=4, sticky=E, padx=5, pady=2)

        Label(labelframe, text='Price:', relief=RIDGE, width=15).grid(row=3, column=1, sticky=W,padx=10, pady=2)
        self.pricefield = Entry(labelframe, width=15)
        self.pricefield.grid(row=3, column=2, sticky=W, padx=5, pady=2)

        Label(labelframe, text='Cleaning:', relief=RIDGE, width=8).grid(row=3, column=3, sticky=E,padx=3, pady=2)
        self.cleaningfield = Entry(labelframe, width=8)
        self.cleaningfield.grid(row=3, column=4, sticky=E, padx=5, pady=2)

        Label(labelframe, text='Airbnb Fee:', relief=RIDGE, width=15).grid(row=4, column=1, sticky=W,padx=10, pady=2)
        self.feefield = Entry(labelframe, width=15)
        self.feefield.grid(row=4, column=2, sticky=W, padx=5, pady=2)

        Label(labelframe, text='Paid:', relief=RIDGE, width=8).grid(row=4, column=3, sticky=E,padx=3, pady=2)
        self.paidfield = Entry(labelframe, width=8)
        self.paidfield.grid(row=4, column=4, sticky=E, padx=5, pady=2)


        ttk.Button(labelframe, text='Save', command=self.on_add_xls_record_button_clicked).grid(
            row=5, column=1, sticky=W, padx=11, pady=2)

        ttk.Button(labelframe, text='Exit', command=lambda: labelframe.grid_remove()).grid(
            row=5, column=2, sticky=W, padx=3, pady=2)

        self.create_left_icon

    def menu_label_frame(self):
        labelframe2 = LabelFrame(self.root, text='Foresthouse Menu')
        labelframe2.grid(row=0, column=0, padx=40, pady=0, sticky='e')

        ttk.Button(labelframe2, text='CalcTax', width=40, command=self.on_calctax_button_clicked).grid(
            row=1, column=1, sticky=W, padx=2, pady=2)

        ttk.Button(labelframe2, text='Activities', width=40, command=self.on_calctax_button_clicked).grid(
            row=2, column=1, sticky=W, padx=2, pady=1)
        ttk.Button(labelframe2, text='Calculator', width=40, command=self.on_calculator_button_clicked).grid(
            row=3, column=1, sticky=W, padx=2, pady=1)
        ttk.Button(labelframe2, text='Accounting', width=40, command=self.on_accounting_button_clicked).grid(
            row=4, column=1, sticky=W, padx=2, pady=1)
        ttk.Button(labelframe2, text='Pictures', width=40, command=self.on_calctax_button_clicked).grid(
            row=5, column=1, sticky=W, padx=2, pady=1)

    def on_add_record_button_clicked(self):
        self.add_new_record()

        self.create_left_icon

    def on_add_xls_record_button_clicked(self):
        self.add_new_xls_record()

        self.create_left_icon
    def on_calctax_button_clicked(self):
        import subprocess
        subprocess.call(["python", "calctax.py"])
        # subprocess.call(["python", "myscript2.py"])

    def on_accounting_button_clicked(self):
        self.create_xls_tree_view()

        #import subprocess
        #subprocess.call(["python", "calctax.py"])
        # subprocess.call(["python", "myscript2.py"])

    def on_calculator_button_clicked(self):
        import subprocess
        subprocess.call(["python", "calculator.py"])
        # subprocess.call(["python", "myscript2.py"])

    def create_tree_view(self):
        self.tree = ttk.Treeview()

        self.tree = ttk.Treeview(height=10, columns=1)
        self.tree.grid(row=4, column=0)

        self.tree.heading('#0', text='ID', anchor=W)
        self.tree.column("#0", width=40)
        style = ttk.Style()

        style.configure("Treeview.Heading", background='blue', foreground='blue')
        self.tree["columns"] = ("one", "two", "three", "four")
        self.tree.column("one", width=150)
        self.tree.column("two", width=100)
        self.tree.column("three", width=200)
        self.tree.column("four", width=250)

        self.tree.heading("one", text="Name", anchor=W)
        self.tree.heading("two", text="Phone", anchor=W)
        self.tree.heading("three", text="Email", anchor=W)
        self.tree.heading("four", text="Comments", anchor=W)

    def create_xls_tree_view(self):
        self.xlstree = ttk.Treeview()

        self.xlstree = ttk.Treeview(height=10, columns=0)
        self.xlstree.grid(row=4, column=0)
        self.xlstree.heading('#0')
        self.xlstree.column("#0", width=0)

        style = ttk.Style()

        style.configure("Treeview.Heading", background='blue', foreground='blue')
        self.xlstree["columns"] = ("","one", "two", "three", "four", "five", "six")

        self.xlstree.column("one", width=150)
        self.xlstree.column("two", width=60)
        self.xlstree.column("three", width=80)
        self.xlstree.column("four", width=80)
        self.xlstree.column("five", width=80)
        self.xlstree.column("six", width=90)

        self.xlstree.heading("", text="Guest/Vendor", anchor=W)
        self.xlstree.heading("one", text="Date", anchor=W,)
        self.xlstree.heading("two", text="Nights", anchor=W)
        self.xlstree.heading("three", text="Price", anchor=W)
        self.xlstree.heading("four", text="Cleaning", anchor=W)
        self.xlstree.heading("five", text="Fees", anchor=W)
        self.xlstree.heading("six", text="Paid", anchor=W)

        wb = load_workbook(filename='airbnb.xlsx')
        ws = wb.active
        row_count = ws.max_row
        column_count = ws.max_column
        items = self.xlstree.get_children()

        for item in items:
            self.xlstree.delete(item)

        # This the original loop when got the warning
        #for row in ws.iter_rows('A2:G10'):
        #    self.tree.insert('', 'end', values=[cell.value for cell in row])

        for row in ws.iter_rows(min_row=1, min_col=1, max_col=7):
            for cell in row:

                if cell.value in ("Guest","Host Date","Nts","Price","Cleaning Fee","Fee","Paid"):
                    break
                else:
                    self.xlstree.insert('', 'end', values=[cell.value for cell in row])

                    break

        self.xlstree.bind('<ButtonRelease-1>', self.select_xls_item)

        self.create_xls_bottom_buttons()

    def select_item(self, a):  # added self and a (event)
        # gets all the values of the selected row
        test_str_library = self.tree.item(self.tree.selection())
        # prints a dictionay of the selected row
        #print ('The test_str = ', type(test_str_library), test_str_library, '\n')
        item = self.tree.selection()[0]  # which row did you click on
        # print ('item clicked ', item) # variable that represents the row you clicked on
        # print (self.tree.item(item)['values'][0]) # prints the first value of the values (the id value)
        record_id = self.tree.item(item)['text']
        # print("Record ID:", record_id)
    def select_xls_item(self, event):

        #print("selected items:")
        for item in self.xlstree.selection():
            item_text = self.xlstree.item(item,"text")
            print(item_text)


    def select_xlsx_item(self, a):  # added self and a (event)
        # gets all the values of the selected row
        test_str_library = self.xlstree.item(self.xlstree.selection())
        # prints a dictionay of the selected row
        #print ('The test_str = ', type(test_str_library), test_str_library, '\n')
        item = self.xlstree.selection()[0]  # which row did you click on
        print(item)
        # print ('item clicked ', item) # variable that represents the row you clicked on
        # print (self.tree.item(item)['values'][0]) # prints the first value of the values (the id value)
        guestfield = self.xlstree.item(item)['text']
        # print("Record ID:", record_id)

    def on_delete_selected_button_clicked(self):
        self.message['text'] = ''
        try:
            # self.tree.item(self.tree.selection())[0]
            self.tree.item(self.tree.selection())['values'][0]
        except IndexError as e:
            self.message['text'] = 'No item selected to delete     '
            return

        self.delete_record()

    def on_delete_xls_selected_button_clicked(self):
        self.message['text'] = ''
        try:
            # self.tree.item(self.tree.selection())[0]
            self.xlstree.item(self.xlstree.selection())['values'][0]
        except IndexError as e:
            self.message['text'] = 'No item selected to delete     '
            return

        self.delete_xls_record()
    def delete_record(self):
        item = self.tree.selection()  # [0]
        idfield = self.tree.item(item)['text']
        conn = sqlite3.connect(self.db_filename)
        c = conn.cursor()
        conn.text_factory = str
        data3 = str(idfield)
        query = "DELETE FROM guest WHERE id = '%s';" % data3
        mydata = c.execute(query)
        conn.commit()
        self.view_records()

    def delete_xls_record(self):

        self.select_xls_item(self.guestfield)
        print('i am in delete_xls_record')
        item = self.xlstree.selection() #[0]

        guestfield = self.xlstree.item(item)['text']
        self.select_xls_item(self.guestfield)
        print(guestfield)
        #conn = sqlite3.connect(self.db_filename)
        #c = conn.cursor()
        #conn.text_factory = str
        data3 = str(guestfield)
        print(item[0])
        #query = "DELETE FROM guest WHERE id = '%s';" % data3
        #mydata = c.execute(query)
        #conn.commit()
        print(data3)

        print('zzzzzzzzzzzzzzzzzzzz')
        self.view_xls_records()

    def on_modify_selected_button_clicked(self):
        self.message['text'] = ''
        try:
            self.tree.item(self.tree.selection())['values'][0]
        except IndexError as e:
            self.message['text'] = 'No item selected to modify     '
            return
        self.open_modify_window()
        # self.show_entry_fields2()
        # self.show_entry_fields1()

    def on_print_selected_button_clicked(self):


        idfield = self.tree.item(self.tree.selection())['text']
        namefield = self.tree.item(self.tree.selection())['values'][0]
        phonefield = self.tree.item(self.tree.selection())['values'][1]
        emailfield = self.tree.item(self.tree.selection())['values'][2]
        commentfield = self.tree.item(self.tree.selection())['values'][3]

        #import subprocess
        #lpr = subprocess.Popen("/usr/bin/lpr", stdin=subprocess.PIPE)
        #lpr.stdin.write("C:\\PyProjects\\learnpy\\foresthouse.py")
        #or
        #import os
        #os.startfile("C:/Users/TestFile.txt", "print")

        import win32ui
        import win32print
        # X from the left margin, Y from top margin
        # both in pixels
        X = 50;
        Y = 50
        #multi_line_string = input_string.split()
        multi_line_string = 'Mike'
        hDC = win32ui.CreateDC()
        hDC.CreatePrinterDC(win32print.GetDefaultPrinter())
        #hDC.CreatePrinterDC(your_printer_name)
        hDC.StartDoc('Microsoft Print to PDF')
        hDC.StartPage()
        #for line in multi_line_string:
        #    hDC.TextOut(X, Y, line)
        #    Y += 100
        #test_str_library = self.tree.item(self.tree.selection())
        # prints a dictionay of the selected row
        # print ('The test_str = ', type(test_str_library), test_str_library, '\n')
        item = self.tree.selection()[0]  # which row did you click on
        #multi_line_string=('The test_str = ', type(test_str_library), test_str_library, '\n')
        multi_line_string=(str(idfield)+'  '+namefield+'  '+phonefield+'  '+emailfield+'  '+commentfield )

        hDC.TextOut(0,0,multi_line_string)
        hDC.EndPage()
        hDC.EndDoc()

    def open_modify_window(self):
        win = Tk()
        Label(win, text="Name").grid(row=0)
        Label(win, text="phone").grid(row=1)
        Label(win, text="Email").grid(row=2)
        Label(win, text="Comment").grid(row=3)

        item = self.tree.selection()  # [0]
        idfield = self.tree.item(item)['text']

        namefield = self.tree.item(self.tree.selection())['values'][0]
        phonefield = self.tree.item(self.tree.selection())['values'][1]
        emailfield = self.tree.item(self.tree.selection())['values'][2]
        commentfield = self.tree.item(self.tree.selection())['values'][3]

        e1 = Entry(win, width=30)
        e2 = Entry(win, width=30)
        e3 = Entry(win, width=30)
        e4 = Entry(win, width=30)

        e1.insert(20, namefield)
        e2.insert(20, phonefield)
        e3.insert(20, emailfield)
        e4.insert(20, commentfield)

        e1.grid(row=0, column=1, padx=10, pady=5)
        e2.grid(row=1, column=1, padx=10, pady=5)
        e3.grid(row=2, column=1, padx=10, pady=5)
        e4.grid(row=3, column=1, padx=10, pady=5)

        Button(win, text='Quit', command=win.destroy, fg='green').grid(row=4, column=1, sticky=W, pady=4, padx=10)
        Button(win, text='Update', fg='red', command=lambda:
        self.modify_record(e1.get(), e2.get(), e3.get(), e4.get())).grid(row=4, column=1, sticky=E, pady=4, padx=10)
        # win.destroy()
        # mainloop()

        # self.windestroy(win)

    def modify_record(self, namefield, phonefield, emailfield, commentfield):
        item = self.tree.selection()  # [0]
        idfield = self.tree.item(item)['text']

        query = 'UPDATE guest SET name=?, phone=?, email=?, comment=? WHERE  id=?'
        parameters = (namefield, phonefield, emailfield, commentfield, idfield)

        with sqlite3.connect(self.db_filename) as conn:
            cursor = conn.cursor()
            #print('parameters=', parameters)
            query_result = cursor.execute(query, parameters)
            # print('query_result=', query_result)
            conn.commit()
        self.message['text'] = 'The record of  {}  was modified     '.format(namefield)
        # self.transient.destroy()
        # self.windestroy(win)
        self.view_records()

        return

    # def windestroy(win):
    #    win.destrroy
    #    #win.quit
    #    return

    def add_update_record(self):
        if self.new_records_validated():
            query = 'INSERT INTO guest VALUES(NULL,?, ?, ?, ?)'
            parameters = (self.namefield.get(), self.phonefield.get(),
                          self.emailfield.get(), self.commentfield.get())
            self.execute_db_query(query, parameters)
            self.message['text'] = 'Phone record of {} {} {} {} added     '.format(
                self.namefield.get(), self.phonefield.get(), self.emailfield.get(), self.commentfield.get())
            self.namefield.delete(0, END)
            self.phonefield.delete(0, END)
            self.emailfield.delete(0, END)
            self.commentfield.delete(0, END)
        else:
            self.message['text'] = 'Name cannot be blank     '
        self.view_records()

    def add_new_record(self):
        if self.new_records_validated():
            query = 'INSERT INTO guest VALUES(NULL,?, ?, ?, ?)'
            parameters = (self.namefield.get(), self.phonefield.get(),
                          self.emailfield.get(), self.commentfield.get())
            # print(parameters)
            self.execute_db_query(query, parameters)
            self.message['text'] = 'New SQL record of {} {} {} {} added     '.format(
                self.namefield.get(), self.phonefield.get(),
                self.emailfield.get(), self.commentfield.get())
            self.namefield.delete(0, END)
            self.phonefield.delete(0, END)
            self.emailfield.delete(0, END)
            self.commentfield.delete(0, END)
        else:
            self.message['text'] = 'Name cannot be blank     '
        self.view_records()

    def add_new_xls_record(self):
        class Foo(float):
            def __new__(cls, value, extra):
                return super().__new__(cls, value)

            def __init__(self, value, extra):
                float.__init__(value)
                self.extra = extra
        if self.new_xls_records_validated():
            style = ttk.Style()

            from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
            #styles = openpyxl.styles()
            #self.hostdatefield.styles.number_format.format_code = 'mm/dd/yyyy'
            parameters = (self.guestfield.get(), self.hostdatefield.get(),
                          self.nitsfield.get(), self.pricefield.get(),self.cleaningfield.get(),
                          self.feefield.get(),self.paidfield.get())
            item = self.tree.selection()  # [0]

            guestfield=self.guestfield.get()
            hostdatefield=self.hostdatefield.get()
            nitsfield=int(self.nitsfield.get())
            pricefield=float(self.pricefield.get())
            cleaningfield=float(self.cleaningfield.get())
            feefield=float(self.feefield.get())
            paidfield=float(self.paidfield.get())
            #self.hostdatefield.style.number_format.format_code = 'mm/dd/yyyy'
            """
            new_date.style.number_format.format_code = 'mm/dd/yyyy'
            wb = openpyxl.load_workbook(file)
            old_sheet = wb.get_sheet_by_name('Sheet1')
            old_sheet.title = 'Sheet1.5'
            max_row = old_sheet.get_highest_row()
            max_col = old_sheet.get_highest_column()
            wb.create_sheet(0, 'Sheet1')

            new_sheet = wb.get_sheet_by_name('Sheet1')

            # Do the header.
            for col_num in range(0, max_col):
            new_sheet.cell(row=0, column=col_num).value = old_sheet.cell(row=0, column=col_num).value

            # The row to be inserted. We're manually populating each cell.
            new_sheet.cell(row=1, column=0).value = 'DUMMY'
            new_sheet.cell(row=1, column=1).value = 'DUMMY'

            """
            wb = load_workbook(filename='airbnb.xlsx')
            ws = wb.active
            row_count = ws.max_row+1
            column_count = ws.max_column
            ws.cell(row=row_count, column=1).value = guestfield
            ws.cell(row=row_count, column=2).value = hostdatefield
            ws.cell(row=row_count, column=3).value = nitsfield
            ws.cell(row=row_count, column=4).value = pricefield
            ws.cell(row=row_count, column=5).value = cleaningfield
            ws.cell(row=row_count, column=6).value = feefield
            ws.cell(row=row_count, column=7).value = paidfield
            #ws.cell(row=row_count, column=7).value = self.paidfield.get()
            wb.save(filename='airbnb.xlsx')
            # print(parameters)
            #self.execute_db_query(query, parameters)

            self.message['text'] = 'New XLS record of {} {} {} {} added     '.format(
                self.guestfield.get(), self.hostdatefield.get(),
                self.nitsfield.get(), self.pricefield.get())
            self.guestfield.delete(0, END)
            self.hostdatefield.delete(0, END)
            self.nitsfield.delete(0, END)
            self.pricefield.delete(0, END)
            self.cleaningfield.delete(0, END)
            self.feefield.delete(0, END)
            self.paidfield.delete(0, END)

        else:
            self.message['text'] = 'Name cannot be blank     '
        self.view_xls_records()

    def create_bottom_buttons(self):
        ttk.Button(text='Add Record',
                   command=self.create_label_frame).grid(row=5, column=0, sticky=W, padx=5)  #W5
        ttk.Button(text='Modify Selected',
                   command=self.on_modify_selected_button_clicked).grid(row=5, column=0, sticky=W, padx=200)  #W200
        ttk.Button(text='Delete Selected',
                   command=self.on_delete_selected_button_clicked).grid(row=5, column=0, sticky=E,padx=200)  #E200
        ttk.Button(text='Print Records',
                   command=self.on_print_selected_button_clicked).grid(row=5, column=0, sticky=E, padx=5)    #E5

    def clear_button(self):
        for button in root.grid_slaves():
            if int(button.grid_info()["row"]) > 4:
               button.grid_forget()

    def create_xls_bottom_buttons(self):
        self.clear_button()

        ttk.Button(text='Add Record',
                   command=self.add_record_xls_label_frame).grid(row=5, column=0, sticky=W, padx=5)
        ttk.Button(text='Modify Selected',
                   command=self.on_modify_selected_button_clicked).grid(row=5, column=0, sticky=W, padx=150)
        ttk.Button(text='Delete Selected',
                   command=self.on_delete_xls_selected_button_clicked).grid(row=5, column=0, sticky=W,padx=320)
        ttk.Button(text='Print Records',
                   command=self.on_print_selected_button_clicked).grid(row=5, column=0, sticky=E, padx=175)
        ttk.Button(text='Exit Accounting',
                   command=self.on_exit_accounting_selected_button_clicked).grid(row=5, column=0, sticky=E, padx=5)

    def on_exit_accounting_selected_button_clicked(self):
        self.clear_button()
        self.create_tree_view()
        self.create_bottom_buttons()
        self.view_records()

    def view_records(self):
        items = self.tree.get_children()
        for item in items:
            self.tree.delete(item)
        query = 'SELECT * FROM guest ORDER BY ID asc'
        guest_table_entries = self.execute_db_query(query)

        for row in guest_table_entries:
            self.tree.insert('', 'end', text=row[0], values=(row[1], row[2], row[3], row[4]))
        self.tree.bind('<ButtonRelease-1>', self.select_item)

    def view_xls_records(self):
        wb = load_workbook(filename='airbnb.xlsx')
        ws = wb.active
        row_count = ws.max_row
        column_count = ws.max_column
        #self.tree = ttk.Treeview()
        items = self.tree.get_children()
        for item in items:
            self.tree.delete(item)

        # This the original loop when got the warning
        # for row in ws.iter_rows('A2:G10'):
        #    self.tree.insert('', 'end', values=[cell.value for cell in row])

        for row in ws.iter_rows(min_row=1, min_col=1, max_col=7):
            for cell in row:
                if cell.value in ("Guest", "Host Date", "Nts", "Price", "Cleaning Fee", "Fee", "Paid"):
                    break
                else:
                    self.tree.insert('', 'end', values=[cell.value for cell in row])
                    break

        self.tree.bind('<ButtonRelease-1>', self.select_xls_item)

    def new_records_validated(self):
        return (self.namefield.get() != 0)  # and len(self.phonefield.get()) != 0

    def new_xls_records_validated(self):
        return (self.guestfield.get() != 0)  # and len(self.phonefield.get()) != 0

    def create_message_area(self):
        self.message = Label(text='', fg='red')
        self.message.grid(row=3, column=0, sticky=E)


if __name__ == '__main__':
    root = Tk()
    application = PhoneBook(root)
    root.mainloop()
